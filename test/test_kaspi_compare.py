import asyncio
import json
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import ANY, AsyncMock, patch

import aiohttp
from openpyxl import load_workbook

from app import (
    build_ozon_products_for_comparison,
    load_ozon_products_from_json,
    run_kaspi_comparison,
)
from services import kaspi_compare
from services.report import REPORT_COLUMNS, save_arbitrage_report


class KaspiQueryAndMatchingTests(unittest.TestCase):
    def test_build_search_query_variants(self):
        self.assertEqual(
            kaspi_compare._build_search_query(
                {
                    "title": "Телефон",
                    "brand": "Apple",
                    "article": "A123",
                }
            ),
            "Apple A123 Телефон",
        )
        self.assertEqual(
            kaspi_compare._build_search_query(
                {"title": "Телефон", "brand": "Apple", "article": None}
            ),
            "Apple Телефон",
        )
        self.assertEqual(
            kaspi_compare._build_search_query(
                {"title": "Телефон", "brand": None, "article": None}
            ),
            "Телефон",
        )

    def test_brand_penalty_and_article_rejection(self):
        ozon = {
            "title": "Apple iPhone 15 128GB",
            "brand": "Apple",
            "article": "A123",
        }
        same_brand = {
            "title": "Apple iPhone 15 128GB",
            "brand": "Apple",
            "article": "A123",
        }
        other_brand = {
            "title": "Apple iPhone 15 128GB",
            "brand": "Samsung",
            "article": "A123",
        }
        other_article = {
            "title": "Apple iPhone 15 128GB",
            "brand": "Apple",
            "article": "B999",
        }

        base_score = kaspi_compare._calculate_match_score(ozon, same_brand)
        penalized = kaspi_compare._calculate_match_score(ozon, other_brand)
        self.assertIsNotNone(base_score)
        self.assertEqual(
            penalized,
            base_score - kaspi_compare.BRAND_MISMATCH_PENALTY,
        )
        self.assertIsNone(
            kaspi_compare._calculate_match_score(ozon, other_article)
        )

    def test_selects_cheapest_candidate_close_to_best_score(self):
        ozon = {"title": "Товар", "price": 10000}
        candidates = [
            {"title": "best", "price": 20000, "url": "best"},
            {"title": "near", "price": 17000, "url": "near"},
            {"title": "far", "price": 15000, "url": "far"},
        ]
        scores = {"best": 90.0, "near": 86.0, "far": 80.0}

        with patch.object(
            kaspi_compare,
            "_calculate_match_score",
            side_effect=lambda _, candidate: scores[candidate["title"]],
        ):
            selected = kaspi_compare._select_candidate(ozon, candidates)

        self.assertIsNotNone(selected)
        self.assertEqual(selected[0]["url"], "near")
        self.assertEqual(selected[1], 86.0)

    def test_rejects_suspicious_price_over_three_times_ozon(self):
        ozon = {"title": "Товар", "price": 10000}
        candidates = [
            {"title": "Товар", "price": 30001, "url": "suspicious"}
        ]
        self.assertIsNone(kaspi_compare._select_candidate(ozon, candidates))

    def test_normalizes_kaspi_card_url_and_sale_price(self):
        candidate = kaspi_compare._normalize_kaspi_card(
            {
                "title": "Товар",
                "unitSalePrice": 19000,
                "unitPrice": 20000,
                "shopLink": "/p/test-123/",
                "brand": "Brand",
            }
        )
        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["price"], 19000.0)
        self.assertEqual(
            candidate["url"],
            "https://kaspi.kz/shop/p/test-123/",
        )


class EconomicsTests(unittest.TestCase):
    def test_delivery_boundary_and_economics(self):
        result = kaspi_compare._calculate_economics(
            {"title": "A", "price": 10000, "url": "ozon"},
            {"title": "A", "price": 20000, "url": "kaspi"},
            90,
        )
        self.assertIsNotNone(result)
        self.assertEqual(result["delivery"], 950.0)
        self.assertEqual(result["total_cost"], 10950.0)
        self.assertEqual(result["net_revenue"], 16900.0)
        self.assertEqual(result["profit"], 5950.0)

        above_boundary = kaspi_compare._calculate_economics(
            {"title": "A", "price": 10001, "url": "ozon"},
            {"title": "A", "price": 22000, "url": "kaspi"},
            90,
        )
        self.assertIsNotNone(above_boundary)
        self.assertEqual(above_boundary["delivery"], 2000.0)

    def test_strict_profit_filter(self):
        result = kaspi_compare._calculate_economics(
            {"title": "A", "price": 10000, "url": "ozon"},
            {
                "title": "A",
                "price": 16508.87573964497,
                "url": "kaspi",
            },
            90,
        )
        self.assertIsNone(result)

    def test_includes_roi_exactly_five_percent(self):
        total_cost = Decimal("10950")
        kaspi_price = (
            total_cost * Decimal("1.05") / Decimal("0.845")
        )
        result = kaspi_compare._calculate_economics(
            {"title": "A", "price": 10000, "url": "ozon"},
            {"title": "A", "price": kaspi_price, "url": "kaspi"},
            90,
            min_roi=Decimal("5"),
            min_profit=Decimal("0"),
        )
        self.assertIsNotNone(result)
        self.assertEqual(result["roi"], 5.0)

    def test_rejects_roi_below_five_percent(self):
        total_cost = Decimal("10950")
        kaspi_price = (
            total_cost * Decimal("1.0499") / Decimal("0.845")
        )
        result = kaspi_compare._calculate_economics(
            {"title": "A", "price": 10000, "url": "ozon"},
            {"title": "A", "price": kaspi_price, "url": "kaspi"},
            90,
            min_roi=Decimal("5"),
            min_profit=Decimal("0"),
        )
        self.assertIsNone(result)

    def test_rejects_profit_exactly_minimum(self):
        total_cost = Decimal("10950")
        kaspi_price = (
            (total_cost + Decimal("3000")) / Decimal("0.845")
        )
        result = kaspi_compare._calculate_economics(
            {"title": "A", "price": 10000, "url": "ozon"},
            {"title": "A", "price": kaspi_price, "url": "kaspi"},
            90,
            min_roi=Decimal("0"),
            min_profit=Decimal("3000"),
        )
        self.assertIsNone(result)


class FakeResponse:
    def __init__(self, payload=None, error=None):
        self.payload = payload
        self.error = error

    async def __aenter__(self):
        if self.error:
            raise self.error
        return self

    async def __aexit__(self, exc_type, exc, traceback):
        return False

    def raise_for_status(self):
        return None

    async def json(self, content_type=None):
        return self.payload


class FakeSession:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = 0

    def get(self, *args, **kwargs):
        self.calls += 1
        return self.responses.pop(0)


class KaspiAsyncTests(unittest.IsolatedAsyncioTestCase):
    async def test_http_retry_then_success(self):
        session = FakeSession(
            [
                FakeResponse(error=aiohttp.ClientConnectionError("one")),
                FakeResponse(error=asyncio.TimeoutError()),
                FakeResponse(payload={"data": {"cards": [{"id": "1"}]}}),
            ]
        )
        with patch("services.kaspi_compare.asyncio.sleep", new=AsyncMock()):
            cards = await kaspi_compare._request_kaspi_page(
                session,
                asyncio.Semaphore(1),
                "test",
                0,
            )
        self.assertEqual(cards, [{"id": "1"}])
        self.assertEqual(session.calls, 3)

    async def test_compare_timeout_returns_empty_report_data(self):
        product = {
            "title": "Test product",
            "price": 10000,
            "url": "ozon-1",
            "brand": None,
            "article": None,
            "category": None,
        }

        async def slow_compare(*args, **kwargs):
            await asyncio.sleep(1)
            return False, None

        with (
            patch(
                "services.kaspi_compare._compare_one",
                side_effect=slow_compare,
            ),
            patch(
                "services.kaspi_compare.COMPARE_TIMEOUT_SECONDS",
                0.01,
            ),
        ):
            results = await kaspi_compare.compare_with_kaspi([product])

        self.assertEqual(results, [])

    async def test_compare_filters_and_sorts_results(self):
        products = [
            {
                "title": "Apple iPhone 15 128GB",
                "price": 10000,
                "url": "ozon-1",
                "brand": "Apple",
                "article": None,
                "category": "phones",
            },
            {
                "title": "Samsung Galaxy S24",
                "price": 10000,
                "url": "ozon-2",
                "brand": "Samsung",
                "article": None,
                "category": "phones",
            },
        ]

        async def fake_search(query):
            if "Apple" in query:
                return [
                    {
                        "title": "Apple iPhone 15 128GB",
                        "price": 22000,
                        "url": "kaspi-1",
                        "brand": "Apple",
                        "article": None,
                    }
                ]
            return [
                {
                    "title": "Samsung Galaxy S24",
                    "price": 20000,
                    "url": "kaspi-2",
                    "brand": "Samsung",
                    "article": None,
                }
            ]

        with patch(
            "services.kaspi_compare.search_kaspi_product",
            side_effect=fake_search,
        ):
            results = await kaspi_compare.compare_with_kaspi(products)

        self.assertEqual(len(results), 2)
        self.assertGreater(results[0]["roi"], results[1]["roi"])
        self.assertEqual(results[0]["kaspi_url"], "kaspi-1")


class ReportAndCliTests(unittest.TestCase):
    def test_report_contains_headers_links_and_empty_rows(self):
        item = {
            "ozon_title": "Ozon item",
            "kaspi_title": "Kaspi item",
            "brand": "Brand",
            "ozon_price": 10000,
            "kaspi_price": 20000,
            "delivery": 950,
            "total_cost": 10950,
            "net_revenue": 16900,
            "profit": 5950,
            "roi": 54.34,
            "match_score": 90,
            "ozon_url": "https://ozon.ru/product/1/",
            "kaspi_url": "https://kaspi.kz/shop/p/1/",
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            with patch(
                "services.report.REPORTS_DIR",
                Path(temp_dir),
            ):
                report_path = save_arbitrage_report([item])
                workbook = load_workbook(report_path)
                sheet = workbook.active

                self.assertEqual(
                    [cell.value for cell in sheet[1]],
                    [title for title, _ in REPORT_COLUMNS],
                )
                self.assertEqual(sheet["A2"].value, 1)
                self.assertEqual(
                    sheet["M2"].hyperlink.target,
                    item["ozon_url"],
                )
                self.assertEqual(
                    sheet["N2"].hyperlink.target,
                    item["kaspi_url"],
                )
                self.assertEqual(sheet.freeze_panes, "A2")

                empty_path = save_arbitrage_report([])
                empty_book = load_workbook(empty_path)
                self.assertEqual(empty_book.active.max_row, 1)

    def test_cli_normalizes_successful_ozon_products(self):
        successful = SimpleNamespace(
            success=True,
            article="ozon-marketplace-id",
            name="Товар",
            price=12345,
            card_price=12000,
        )
        failed = SimpleNamespace(
            success=False,
            article="failed",
            name="Ошибка",
            price=1,
            card_price=1,
        )
        manager = SimpleNamespace(
            last_results={
                "products": [successful, failed],
                "links": {
                    "https://ozon.ru/product/test-ozon-marketplace-id/": ""
                },
            }
        )

        result = build_ozon_products_for_comparison(manager)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["price"], 12345)
        self.assertIsNone(result[0]["article"])
        self.assertIsNone(result[0]["brand"])
        self.assertIn("ozon-marketplace-id", result[0]["url"])

    def test_loads_all_thirty_successful_products_from_json(self):
        products = [
            {
                "article": str(index),
                "name": f"Товар {index}",
                "price": 10000 + index,
                "product_url": f"https://ozon.ru/product/{index}/",
                "success": True,
            }
            for index in range(30)
        ]
        products.append(
            {
                "article": "failed",
                "name": "Ошибка",
                "price": 1,
                "product_url": "",
                "success": False,
            }
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            json_path = Path(temp_dir) / "ozon.json"
            json_path.write_text(
                json.dumps(
                    {
                        "category_url": "https://ozon.kz/category/test/",
                        "products": products,
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            result = load_ozon_products_from_json(json_path)

        self.assertEqual(len(result), 30)
        self.assertTrue(all(item["article"] is None for item in result))
        self.assertTrue(
            all(
                item["category"] == "https://ozon.kz/category/test/"
                for item in result
            )
        )

    def test_cli_comparison_calls_services(self):
        product = SimpleNamespace(
            success=True,
            article="ozon-id",
            name="Товар",
            price=10000,
            card_price=10000,
        )
        manager = SimpleNamespace(
            last_results={
                "products": [product],
                "links": {"https://ozon.ru/product/test-ozon-id/": ""},
            }
        )
        report_items = [{"roi": 40, "profit": 5000}]

        with patch(
            "services.kaspi_compare.compare_with_kaspi",
            new=AsyncMock(return_value=report_items),
        ) as compare_mock, patch(
            "services.report.save_arbitrage_report",
            return_value="reports/test.xlsx",
        ) as report_mock:
            count, path = run_kaspi_comparison(manager)

        self.assertEqual(count, 1)
        self.assertEqual(path, "reports/test.xlsx")
        self.assertEqual(compare_mock.await_count, 1)
        compare_mock.assert_awaited_once_with(
            ANY,
            min_roi=25.0,
            min_profit=3000.0,
        )
        report_mock.assert_called_once_with(report_items)


if __name__ == "__main__":
    unittest.main()
