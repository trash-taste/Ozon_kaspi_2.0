import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
REPORTS_DIR = Path(__file__).resolve().parent.parent / "reports"
MONEY_QUANT = Decimal("0.01")
FEE_RATE = Decimal("0.16")

REPORT_COLUMNS = [
    ("№", None),
    ("Товар", "product_title"),
    ("Цена Ozon", "ozon_price"),
    ("Цена продажи", "sale_price"),
    ("Откуда цена", "source"),
    ("Чистыми с Kaspi", "net_from_kaspi"),
    ("Прибыль", "profit"),
    ("ROI %", "roi"),
    ("Ссылка Ozon", "ozon_url"),
    ("Ссылка источника цены", "source_url"),
]

MONEY_COLUMNS = {3, 4, 6, 7}
LINK_COLUMNS = {9, 10}
ROI_COLUMN = 8

INTERNET_REPORT_COLUMNS = [
    *REPORT_COLUMNS,
]

INTERNET_MONEY_COLUMNS = MONEY_COLUMNS
INTERNET_LINK_COLUMNS = LINK_COLUMNS


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    normalized = str(value).replace("\u00a0", "").replace(" ", "")
    normalized = normalized.replace(",", ".")
    normalized = "".join(
        char for char in normalized if char.isdigit() or char in ".-"
    )
    if not normalized:
        return None
    try:
        result = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None
    if not result.is_finite():
        return None
    return result


def _money(value: Decimal) -> float:
    return float(value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP))


def _delivery_for(ozon_price: Decimal) -> Decimal:
    return Decimal("950") if ozon_price <= Decimal("10000") else Decimal("2000")


def _calculate_decision_financials(
    ozon_price: Decimal,
    sale_price: Decimal,
    fee_rate: Decimal = FEE_RATE,
) -> dict[str, float]:
    delivery = _delivery_for(ozon_price)
    net_from_kaspi = sale_price * (Decimal("1") - fee_rate) - delivery
    profit = net_from_kaspi - ozon_price
    roi = profit / ozon_price * Decimal("100")
    return {
        "net_from_kaspi": _money(net_from_kaspi),
        "profit": _money(profit),
        "roi": _money(roi),
    }


def _source_label(source: Any, source_url: str = "", default: str = "Other") -> str:
    raw = str(source or source_url or "").casefold()
    host = urlparse(source_url or raw).netloc.casefold().removeprefix("www.")
    value = host or raw
    if "kaspi" in value:
        return "Kaspi"
    if "sulpak" in value:
        return "Sulpak"
    if "technodom" in value:
        return "Technodom"
    if "mechta" in value:
        return "Mechta"
    if "shop.kz" in value:
        return "Shop.kz"
    if "flip" in value:
        return "Flip"
    if "wildberries" in value or "wb." in value:
        return "Wildberries"
    return default


def _decision_rows(
    items: list[dict],
    *,
    default_source: str,
    sale_price_key: str,
    source_url_key: str,
) -> list[dict[str, Any]]:
    rows = []
    for item in items:
        ozon_price = _to_decimal(item.get("ozon_price"))
        sale_price = _to_decimal(item.get(sale_price_key))
        if ozon_price is not None and ozon_price <= 0:
            ozon_price = None

        product_title = str(
            item.get("ozon_title")
            or item.get("title")
            or item.get("internet_title")
            or item.get("kaspi_title")
            or ""
        )
        ozon_url = str(item.get("ozon_url") or "")
        if not product_title and not ozon_url:
            continue

        source_url = str(item.get(source_url_key) or "")
        source = ""
        if sale_price is not None:
            source = _source_label(
                item.get("source"),
                source_url,
                default=default_source,
            )
        fee_rate = _to_decimal(item.get("commission_rate"))
        if fee_rate is None:
            fee_rate = FEE_RATE
        elif fee_rate > 1:
            fee_rate = fee_rate / Decimal("100")

        row = {
            "product_title": product_title,
            "ozon_price": _money(ozon_price) if ozon_price is not None else None,
            "sale_price": _money(sale_price) if sale_price is not None else None,
            "source": source,
            "ozon_url": ozon_url,
            "source_url": source_url,
            "net_from_kaspi": None,
            "profit": None,
            "roi": None,
        }
        if ozon_price is not None and sale_price is not None:
            row.update(
                _calculate_decision_financials(
                    ozon_price,
                    sale_price,
                    fee_rate=fee_rate,
                )
            )
        rows.append(row)

    rows.sort(
        key=lambda row: (row.get("roi") or -10**9, row.get("profit") or -10**9),
        reverse=True,
    )
    return rows


def _write_decision_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
    *,
    title: str,
    header_color: str,
) -> None:
    sheet = workbook.active
    sheet.title = title
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(
        fill_type="solid",
        start_color=header_color,
        end_color=header_color,
    )
    header_font = Font(color="FFFFFF", bold=True)
    for column, (header, _) in enumerate(REPORT_COLUMNS, 1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, item in enumerate(rows, 2):
        sheet.cell(row=row_index, column=1, value=row_index - 1)
        for column, (_, key) in enumerate(REPORT_COLUMNS[1:], 2):
            value = item.get(key) if key else None
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in MONEY_COLUMNS and value is not None:
                cell.number_format = (
                    '#,##0 "₸"'
                    if column in {3, 4}
                    else '#,##0.00 "₸"'
                )
            elif column == ROI_COLUMN and value is not None:
                cell.number_format = '0.00"%"'
            elif column in LINK_COLUMNS and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"

            if key == "source" and value and value != "Kaspi":
                cell.comment = Comment(
                    "Нужна ручная проверка: источник цены не Kaspi.",
                    "OzonParser",
                )

    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:J{last_row}"
    sheet.row_dimensions[1].height = 24
    _autofit_columns(sheet)


def _autofit_columns(sheet) -> None:
    max_widths = {
        1: 8,
        2: 60,
        9: 55,
        10: 55,
    }
    for column in range(1, sheet.max_column + 1):
        max_length = 0
        for cell in sheet.iter_cols(
            min_col=column,
            max_col=column,
            min_row=1,
            max_row=sheet.max_row,
        ):
            for value_cell in cell:
                value = value_cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
        width = min(max(max_length + 2, 10), max_widths.get(column, 22))
        sheet.column_dimensions[get_column_letter(column)].width = width


def save_arbitrage_report(items: list[dict]) -> str:
    """Сохраняет результаты арбитража в Excel и возвращает путь."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = datetime.now().strftime("arbitrage_%Y_%m_%d_%H_%M.xlsx")
    report_path = REPORTS_DIR / filename

    workbook = Workbook()
    rows = _decision_rows(
        items,
        default_source="Kaspi",
        sale_price_key="kaspi_price",
        source_url_key="kaspi_url",
    )
    _write_decision_sheet(
        workbook,
        rows,
        title="Arbitrage",
        header_color="1F4E78",
    )

    workbook.save(report_path)
    logger.info("Excel-отчет арбитража сохранен: %s", report_path)
    return str(report_path)


def save_internet_comparison_report(items: list[dict]) -> str:
    """Сохраняет сравнение Ozon с интернет-ценами в Excel."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = datetime.now().strftime(
        "internet_comparison_%Y_%m_%d_%H_%M.xlsx"
    )
    report_path = REPORTS_DIR / filename

    workbook = Workbook()
    rows = _decision_rows(
        items,
        default_source="Other",
        sale_price_key="internet_price",
        source_url_key="internet_url",
    )
    _write_decision_sheet(
        workbook,
        rows,
        title="Ozon vs Internet",
        header_color="E85D04",
    )

    workbook.save(report_path)
    logger.info(
        "Excel-отчет интернет-сравнения сохранен: %s",
        report_path,
    )
    return str(report_path)
